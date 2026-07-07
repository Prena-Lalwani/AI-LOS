"""
reports_analytics.py — Reports & Analytics module layer.

This module does NOT generate data or train models. It aggregates and
re-presents outputs already produced by the Executive, Demand, Inventory,
Dispatch and Fleet modules. main.py builds those payloads ONCE at startup and
hands them to load_modules() here, so every report is a repackaging of
already-cached results — nothing is recomputed per request.

  * Executive data has no backend model (it is precomputed on the frontend); we
    read the same computed artifact, src/data/execIntelligenceData.json.
  * The other four come straight from their cached FastAPI payloads.

What it exposes:
  - REPORTS registry (6 reports) + list_reports() for GET /api/reports/list
  - build_detail(report_id, date_from, date_to) for GET /api/reports/{id},
    returning a generic { sections: [...] } payload (kpis / table / chart
    blocks) the frontend renders uniformly. Time-series sections honour the
    optional date_from / date_to filter.
  - export_pdf() / export_excel() render the report's tables into a real
    downloadable file (reportlab / openpyxl).

The "Cross-Module KPI Summary" is the one piece of genuinely new logic — it
pulls one or two headline KPIs from each of the five modules into one view.

Scheduling is illustrative only (static next_scheduled = today + a fixed
interval); there is no live scheduler.
"""

import json
import os
from datetime import datetime, timedelta
from io import BytesIO

import pandas as pd

# --------------------------------------------------------------------------- #
# Cached module payloads (populated by main.py at startup via load_modules)   #
# --------------------------------------------------------------------------- #
_CACHE = {}

# Executive is precomputed (it has no backend model). Prefer a backend-local copy
# so the deployed backend is self-contained (the frontend src/ folder is NOT part
# of the backend deploy on Render/Railway); fall back to the frontend source when
# running locally from the repo.
_EXEC_CANDIDATES = [
    os.path.join(os.path.dirname(__file__), "..", "data", "executive", "execIntelligenceData.json"),
    os.path.join(os.path.dirname(__file__), "..", "..", "src", "data", "execIntelligenceData.json"),
]


def _load_executive():
    for path in _EXEC_CANDIDATES:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    raise FileNotFoundError(
        "execIntelligenceData.json not found in: " + " | ".join(os.path.normpath(p) for p in _EXEC_CANDIDATES))


def load_modules(demand=None, inventory=None, dispatch=None, fleet=None):
    """Called once at startup with the already-built payloads. Executive is read
    from its precomputed JSON artifact (it has no backend model)."""
    if demand is not None:
        _CACHE["demand"] = demand
    if inventory is not None:
        _CACHE["inventory"] = inventory
    if dispatch is not None:
        _CACHE["dispatch"] = dispatch
    if fleet is not None:
        _CACHE["fleet"] = fleet
    _CACHE["executive"] = _load_executive()
    return _CACHE


def _module(name):
    """Return a cached module payload, lazily building it as a safety net if the
    endpoint is hit before startup wiring ran (normally always cached)."""
    if name not in _CACHE:
        if name == "executive":
            _CACHE["executive"] = _load_executive()
        elif name == "demand":
            from models.demand_forecast import build_payload
            _CACHE["demand"] = build_payload()
        elif name == "inventory":
            from models.inventory_intelligence import build_payload as b
            _CACHE["inventory"] = b()
        elif name == "dispatch":
            from models.dispatch_intelligence import build_plan
            _CACHE["dispatch"] = build_plan()
        elif name == "fleet":
            from models.fleet_intelligence import build_payload as b
            _CACHE["fleet"] = b()
    return _CACHE[name]


# --------------------------------------------------------------------------- #
# Report registry                                                             #
# --------------------------------------------------------------------------- #
# interval = illustrative days until the next (mock) scheduled run.
REPORTS = [
    {"id": "executive-summary", "title": "Executive Summary", "module": "Executive",
     "description": "Monthly headline KPIs, revenue vs profit trend and top regions.",
     "interval": 30},
    {"id": "demand-forecast", "title": "Demand & Forecast Report", "module": "Demand",
     "description": "Weekly demand trend, Prophet 4-week forecast and seasonal mix.",
     "interval": 7},
    {"id": "inventory-health", "title": "Inventory Health Report", "module": "Inventory",
     "description": "Reorder recommendations, stock-out risk and overstock exposure.",
     "interval": 1},
    {"id": "dispatch-performance", "title": "Dispatch Performance Report", "module": "Dispatch",
     "description": "Route assignment, load balancing and delay / overtime risk.",
     "interval": 1},
    {"id": "fleet-maintenance", "title": "Fleet Maintenance Report", "module": "Fleet",
     "description": "Breakdown risk, driver performance and fuel-efficiency trend.",
     "interval": 7},
    {"id": "cross-module-kpi", "title": "Cross-Module KPI Summary", "module": "All",
     "description": "Headline KPIs pulled from all five modules into one view.",
     "interval": 1},
]
_BY_ID = {r["id"]: r for r in REPORTS}
REPORT_IDS = [r["id"] for r in REPORTS]

_SCHEDULE_NOTE = "Illustrative only — no live scheduler is running."


def _today():
    return datetime.now().date()


def list_reports():
    today = _today()
    out = []
    for r in REPORTS:
        nxt = today + timedelta(days=r["interval"])
        out.append({
            "id": r["id"], "title": r["title"], "module": r["module"],
            "description": r["description"],
            "lastGenerated": today.strftime("%Y-%m-%d"),
            "nextScheduled": nxt.strftime("%Y-%m-%d"),
            "scheduleIntervalDays": r["interval"],
            "scheduleNote": _SCHEDULE_NOTE,
        })
    return out


# --------------------------------------------------------------------------- #
# Formatting + section helpers                                                #
# --------------------------------------------------------------------------- #
def _usd(v):
    v = float(v)
    if abs(v) >= 1_000_000:
        return f"${v / 1_000_000:.1f}M"
    if abs(v) >= 1_000:
        return f"${v / 1_000:.0f}K"
    return f"${v:.0f}"


def _fmt_exec_value(value, unit):
    if unit == "$":
        return _usd(value)
    if unit == "%":
        return f"{value}%"
    return f"{value:,.0f}" if isinstance(value, (int, float)) else str(value)


def _fmt_delta(delta, unit):
    sign = "+" if delta > 0 else ""
    suffix = " pts" if unit == "%" else "%"
    return f"{sign}{delta}{suffix} vs prior"


def _delta_state(delta):
    return "flow" if delta > 0 else "attention" if delta < 0 else "neutral"


def _fmt_num_unit(value, unit):
    if unit == "%":
        return f"{value}%"
    if isinstance(value, (int, float)):
        return f"{value:,.0f}"
    return str(value)


def _kpis(title, items):
    return {"type": "kpis", "title": title, "items": items}


def _table(title, columns, rows, note=None):
    return {"type": "table", "title": title, "columns": columns, "rows": rows, "note": note}


def _chart(title, data, x_key, date_key, series, y_prefix="", y_suffix="", y_thousands=False):
    return {"type": "chart", "title": title, "data": data, "xKey": x_key,
            "dateKey": date_key, "series": series,
            "yPrefix": y_prefix, "ySuffix": y_suffix, "yThousands": y_thousands}


def _filter_by_date(rows, key, dfrom, dto):
    """Filter already-computed time-series rows to [dfrom, dto] (inclusive).
    Month strings like '2024-07' are treated as that month's first day."""
    if not dfrom and not dto:
        return rows
    lo = pd.Timestamp(dfrom) if dfrom else None
    hi = pd.Timestamp(dto) if dto else None
    out = []
    for r in rows:
        try:
            ts = pd.Timestamp(r[key])
        except Exception:
            out.append(r)
            continue
        if lo is not None and ts < lo:
            continue
        if hi is not None and ts > hi:
            continue
        out.append(r)
    return out


def _kpi_by_label(payload, label):
    for k in payload.get("kpis", []):
        if k.get("label") == label:
            return k
    return None


# --------------------------------------------------------------------------- #
# Per-report detail builders                                                  #
# --------------------------------------------------------------------------- #
def _detail_executive(dfrom, dto):
    ex = _module("executive")
    kpi_items = [{
        "label": k["label"], "value": _fmt_exec_value(k["value"], k["unit"]),
        "delta": _fmt_delta(k["delta"], k["unit"]), "state": _delta_state(k["delta"]),
    } for k in ex["kpis"]]

    trend = _filter_by_date(ex["trend"], "month", dfrom, dto)
    chart = _chart(
        "Revenue vs Profit (monthly)",
        [{"month": t["month"], "revenue": round(t["revenue"] / 1e6, 2),
          "profit": round(t["profit"] / 1e6, 2)} for t in trend],
        "month", "month",
        [{"key": "revenue", "label": "Revenue", "color": "var(--accent-flow)", "type": "line"},
         {"key": "profit", "label": "Profit", "color": "var(--accent-attention)", "type": "line"}],
        y_prefix="$", y_suffix="M",
    )

    regions = sorted(ex["regions"], key=lambda r: -r["sales"])[:10]
    region_tbl = _table(
        "Top Regions by Sales",
        [{"key": "region", "label": "Region"}, {"key": "sales", "label": "Sales"},
         {"key": "profit", "label": "Profit"}, {"key": "onTime", "label": "On-Time"}],
        [{"region": r["region"], "sales": _usd(r["sales"]),
          "profit": _usd(r["profit"]), "onTime": f"{r['onTimePct']}%"} for r in regions],
    )
    return {"timeSeries": True, "sections": [
        _kpis("Headline KPIs", kpi_items), chart, region_tbl]}


def _detail_demand(dfrom, dto):
    dm = _module("demand")
    kpi_items = [{
        "label": k["label"], "value": _fmt_num_unit(k["value"], k.get("unit", "")),
        "delta": None, "state": "neutral",
    } for k in dm["kpis"]]

    weekly = _filter_by_date(dm["weeklyTrend"], "weekStart", dfrom, dto)
    chart = _chart(
        "Weekly Demand (units)",
        [{"weekStart": w["weekStart"], "actual": round(w["actual"])} for w in weekly],
        "weekStart", "weekStart",
        [{"key": "actual", "label": "Actual", "color": "var(--accent-flow)", "type": "line"}],
        y_thousands=True,
    )

    forecast_tbl = _table(
        "Prophet Forecast · next 4 weeks",
        [{"key": "weekStart", "label": "Week of"}, {"key": "projected", "label": "Projected"},
         {"key": "lower", "label": "Low"}, {"key": "upper", "label": "High"}],
        [{"weekStart": f["weekStart"], "projected": f"{round(f['projected']):,}",
          "lower": f"{round(f['lower']):,}", "upper": f"{round(f['upper']):,}"}
         for f in dm["forecast"]],
    )
    seasonal_tbl = _table(
        "Seasonal Demand Mix",
        [{"key": "season", "label": "Season"}, {"key": "share", "label": "Share"},
         {"key": "avgDaily", "label": "Avg Units/Day"}],
        [{"season": s["season"], "share": f"{s['sharePct']}%",
          "avgDaily": f"{round(s['avgDailyUnits']):,}"} for s in dm["seasonalBreakdown"]],
    )
    return {"timeSeries": True, "sections": [
        _kpis("Headline KPIs", kpi_items), chart, forecast_tbl, seasonal_tbl]}


def _detail_inventory(dfrom, dto):
    inv = _module("inventory")
    so = inv["stockOutPredictions"]
    kpi_items = [
        {"label": "Reorder Flags", "value": str(inv["reorderRecommendations"]["flaggedCount"]),
         "delta": None, "state": "attention"},
        {"label": "Stock-Out Risks", "value": str(so["atRiskCount"]),
         "delta": "next 7 days", "state": "critical"},
        {"label": "Overstock Value", "value": _usd(inv["overstock"]["totalExcessValue"]),
         "delta": f"{inv['overstock']['count']} locations", "state": "attention"},
        {"label": "Stock-Out Model Acc", "value": f"{so['modelMetrics']['accuracy']}%",
         "delta": "RandomForest", "state": "flow"},
    ]

    reorder_tbl = _table(
        "Reorder Recommendations",
        [{"key": "name", "label": "Product"}, {"key": "wh", "label": "WH"},
         {"key": "onHand", "label": "On Hand"}, {"key": "rop", "label": "ROP"},
         {"key": "qty", "label": "Rec Qty"}, {"key": "status", "label": "Status"}],
        [{"name": r["name"], "wh": r["warehouseId"], "onHand": f"{r['onHand']:,}",
          "rop": f"{r['reorderPoint']:,}", "qty": f"{r.get('recommendedQty') or 0:,}",
          "status": r["status"]} for r in inv["reorderRecommendations"]["rows"][:12]],
    )
    risk_tbl = _table(
        "Stock-Out Risk · next 7 days",
        [{"key": "name", "label": "Product"}, {"key": "wh", "label": "WH"},
         {"key": "onHand", "label": "On Hand"}, {"key": "risk", "label": "Risk"},
         {"key": "days", "label": "Days Left"}],
        [{"name": s["name"], "wh": s["warehouseId"], "onHand": f"{s['onHand']:,}",
          "risk": f"{round(s['riskProbability'] * 100)}%",
          "days": f"{s['daysToStockout']}" if s.get("daysToStockout") is not None else "imminent"}
         for s in so["atRisk"][:12]],
    )
    over_tbl = _table(
        "Overstock Exposure",
        [{"key": "name", "label": "Product"}, {"key": "wh", "label": "WH"},
         {"key": "excess", "label": "Excess Value"}],
        [{"name": o["name"], "wh": o.get("warehouseId", "—"), "excess": _usd(o["excessValue"])}
         for o in inv["overstock"]["rows"][:12]],
    )
    return {"timeSeries": False, "sections": [
        _kpis("Headline KPIs", kpi_items), reorder_tbl, risk_tbl, over_tbl]}


def _detail_dispatch(dfrom, dto):
    dp = _module("dispatch")
    kpi_items = [{
        "label": k["label"], "value": str(k["value"]),
        "delta": k.get("delta"), "state": k.get("state", "neutral"),
    } for k in dp["kpis"]]

    def _route_status(r):
        if r["overtimeRisk"]:
            return "OT risk"
        if r["rerouteRecommended"]:
            return "Reroute"
        if r["capacityUtilizationPct"] < 50:
            return "Underused"
        return "Optimized"

    routes_tbl = _table(
        f"Route Assignment · {dp['warehouseName']} · {dp['planDate']}",
        [{"key": "vehicleId", "label": "Truck"}, {"key": "driver", "label": "Driver"},
         {"key": "stops", "label": "Stops"}, {"key": "dist", "label": "Distance"},
         {"key": "load", "label": "Load"}, {"key": "status", "label": "Status"}],
        [{"vehicleId": r["vehicleId"], "driver": r["driver"], "stops": r["numStops"],
          "dist": f"{r['distanceKm']} km", "load": f"{r['capacityUtilizationPct']}%",
          "status": _route_status(r)} for r in dp["routes"]],
    )

    lb = dp["loadBalancing"]
    lb_items = [
        {"label": "Avg Load Factor", "value": f"{lb['avgUtilizationPct']}%",
         "delta": None, "state": "flow" if lb["avgUtilizationPct"] >= 70 else "attention"},
        {"label": "Underused Routes", "value": str(len(lb["underused"])),
         "delta": "< 50% capacity", "state": "attention" if lb["underused"] else "flow"},
        {"label": "Overloaded Routes", "value": str(len(lb["overloaded"])),
         "delta": ">= 100% capacity", "state": "critical" if lb["overloaded"] else "flow"},
        {"label": "Delay Model MAE", "value": f"{dp['delayRisk']['modelMae']} min",
         "delta": "RandomForest", "state": "flow"},
    ]

    flags = dp["delayRisk"]["flags"]
    delay_tbl = _table(
        "Delay & Overtime Flags",
        [{"key": "vehicleId", "label": "Truck"}, {"key": "driver", "label": "Driver"},
         {"key": "kind", "label": "Flag"}, {"key": "detail", "label": "Detail"}],
        ([{"vehicleId": f["vehicleId"], "driver": f["driver"], "kind": "Reroute",
           "detail": f"+{f['overshootMin']} min vs planned"} for f in flags]
         + [{"vehicleId": f["vehicleId"], "driver": f["driver"], "kind": "Overtime",
             "detail": f"+{f['overMin']} min over shift"} for f in dp["overtimeFlags"]])
        or [{"vehicleId": "—", "driver": "—", "kind": "None", "detail": "All routes within limits"}],
    )
    return {"timeSeries": False, "sections": [
        _kpis("Headline KPIs", kpi_items), routes_tbl,
        _kpis("Load Balancing", lb_items), delay_tbl]}


def _detail_fleet(dfrom, dto):
    fl = _module("fleet")
    m = fl["maintenanceRisk"]["modelMetrics"]
    fa = fl["fuelAnalytics"]
    kpi_items = [
        {"label": "High-Risk Vehicles", "value": str(fl["maintenanceRisk"]["highRiskCount"]),
         "delta": "breakdown < 10d", "state": "critical"},
        {"label": "Breakdown Model Acc", "value": f"{m['accuracy']}%",
         "delta": f"prec {m['precision']}% · rec {m['recall']}%", "state": "flow"},
        {"label": "Fleet Fuel Economy", "value": f"{fa['fleetAvgKmL']} km/L",
         "delta": f"{len(fa['degradingVehicles'])} degrading", "state": "attention"},
        {"label": "Anomalous Vehicles", "value": str(fl["anomalies"]["summary"]["vehiclesFlagged"]),
         "delta": f"last {fl['anomalies']['summary']['windowDays']}d", "state": "attention"},
    ]

    risk_tbl = _table(
        "Predictive Maintenance Risk",
        [{"key": "vehicleId", "label": "Vehicle"}, {"key": "risk", "label": "Risk"},
         {"key": "level", "label": "Level"}, {"key": "brake", "label": "Brake Wear"},
         {"key": "oil", "label": "Oil Δ7d"}],
        [{"vehicleId": v["vehicleId"], "risk": f"{v['riskPct']}%", "level": v["riskLevel"],
          "brake": f"{v['brakePadWearPct']}%", "oil": f"{v['oilPressureTrend7']}"}
         for v in fl["maintenanceRisk"]["vehicles"][:12]],
    )
    drv = fl["driverPerformance"]["drivers"]
    driver_tbl = _table(
        "Driver Performance (top & bottom)",
        [{"key": "rank", "label": "#"}, {"key": "name", "label": "Driver"},
         {"key": "kmL", "label": "km/L"}, {"key": "safety", "label": "Safety"},
         {"key": "harsh", "label": "Harsh/100km"}],
        [{"rank": d["rank"], "name": d["name"], "kmL": d["fuelEfficiencyKmL"],
          "safety": d["safetyScore"], "harsh": d["harshPer100km"]}
         for d in (drv[:5] + drv[-5:])],
    )
    weeks = _filter_by_date(fa["fuelWeeks"], "date", dfrom, dto)
    chart = _chart(
        "Fleet Fuel Efficiency (km/L)",
        [{"label": w["label"], "kmL": w["kmL"]} for w in weeks],
        "label", "date",
        [{"key": "kmL", "label": "km/L", "color": "var(--accent-flow)", "type": "bar"}],
    )
    return {"timeSeries": True, "sections": [
        _kpis("Headline KPIs", kpi_items), risk_tbl, driver_tbl, chart]}


def _detail_cross_module(dfrom, dto):
    """The one piece of genuinely new logic: one or two headline KPIs pulled from
    each of the five modules into a single combined view."""
    ex = _module("executive")
    dm = _module("demand")
    inv = _module("inventory")
    dp = _module("dispatch")
    fl = _module("fleet")

    rev = _kpi_by_label({"kpis": ex["kpis"]}, "Total Revenue")
    avg_ontime = round(sum(r["onTimePct"] for r in ex["regions"]) / len(ex["regions"]), 1)
    fc_acc = _kpi_by_label(dm, "Forecast Accuracy")
    load = _kpi_by_label(dp, "Avg Load Factor")

    combined = [
        {"module": "Executive", "metric": "Total Revenue",
         "value": _fmt_exec_value(rev["value"], rev["unit"]) if rev else "—", "state": "flow"},
        {"module": "Executive", "metric": "Avg Region On-Time",
         "value": f"{avg_ontime}%", "state": "flow" if avg_ontime >= 90 else "attention"},
        {"module": "Demand", "metric": "Forecast Accuracy",
         "value": f"{fc_acc['value']}%" if fc_acc else "—", "state": "flow"},
        {"module": "Inventory", "metric": "Stock-Out Risks (7d)",
         "value": str(inv["stockOutPredictions"]["atRiskCount"]), "state": "critical"},
        {"module": "Dispatch", "metric": "Avg Load Factor",
         "value": str(load["value"]) if load else "—", "state": "flow"},
        {"module": "Fleet", "metric": "High-Risk Vehicles",
         "value": str(fl["maintenanceRisk"]["highRiskCount"]), "state": "critical"},
        {"module": "Fleet", "metric": "Breakdown Model Accuracy",
         "value": f"{fl['maintenanceRisk']['modelMetrics']['accuracy']}%", "state": "flow"},
    ]

    kpi_items = [{"label": f"{c['module']} · {c['metric']}", "value": c["value"],
                  "delta": None, "state": c["state"]} for c in combined]
    combined_tbl = _table(
        "Cross-Module Headline KPIs",
        [{"key": "module", "label": "Module"}, {"key": "metric", "label": "Metric"},
         {"key": "value", "label": "Value"}],
        [{"module": c["module"], "metric": c["metric"], "value": c["value"]} for c in combined],
        note="One combined view drawing the headline number from each module.",
    )
    return {"timeSeries": False, "sections": [
        _kpis("All Modules at a Glance", kpi_items), combined_tbl]}


_BUILDERS = {
    "executive-summary": _detail_executive,
    "demand-forecast": _detail_demand,
    "inventory-health": _detail_inventory,
    "dispatch-performance": _detail_dispatch,
    "fleet-maintenance": _detail_fleet,
    "cross-module-kpi": _detail_cross_module,
}


def build_detail(report_id, date_from=None, date_to=None):
    """Full detail payload for one report. Raises KeyError for an unknown id."""
    meta = _BY_ID[report_id]                       # KeyError -> 404 in main.py
    body = _BUILDERS[report_id](date_from, date_to)
    today = _today()
    return {
        "id": meta["id"], "title": meta["title"], "module": meta["module"],
        "description": meta["description"],
        "lastGenerated": today.strftime("%Y-%m-%d"),
        "nextScheduled": (today + timedelta(days=meta["interval"])).strftime("%Y-%m-%d"),
        "scheduleNote": _SCHEDULE_NOTE,
        "timeSeries": body["timeSeries"],
        "dateFilter": {"from": date_from, "to": date_to},
        "sections": body["sections"],
    }


# --------------------------------------------------------------------------- #
# Exports — flatten a report's sections into simple header+row tables          #
# --------------------------------------------------------------------------- #
def _section_table(section):
    """Return (headers, rows) for any section type — the export representation."""
    if section["type"] == "kpis":
        return (["Metric", "Value"], [[it["label"], str(it["value"])] for it in section["items"]])
    if section["type"] == "table":
        cols = section["columns"]
        return ([c["label"] for c in cols],
                [[str(r.get(c["key"], "")) for c in cols] for r in section["rows"]])
    if section["type"] == "chart":
        keys = [section["xKey"]] + [s["key"] for s in section["series"]]
        headers = [section["xKey"]] + [s["label"] for s in section["series"]]
        return (headers, [[str(row.get(k, "")) for k in keys] for row in section["data"]])
    return ([], [])


def export_pdf(report_id, date_from=None, date_to=None):
    """Render the report's tables into a simple, clean PDF. Returns raw bytes."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle)

    detail = build_detail(report_id, date_from, date_to)
    styles = getSampleStyleSheet()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, title=detail["title"],
                            leftMargin=0.7 * inch, rightMargin=0.7 * inch,
                            topMargin=0.7 * inch, bottomMargin=0.7 * inch)
    avail_w = letter[0] - 1.4 * inch
    elems = [Paragraph(detail["title"], styles["Title"])]
    sub = f"{detail['module']} module · generated {detail['lastGenerated']}"
    if date_from or date_to:
        sub += f" · filtered {date_from or '…'} → {date_to or '…'}"
    elems += [Paragraph(sub, styles["Normal"]), Spacer(1, 14)]

    header_bg = colors.HexColor("#1b242c")
    for section in detail["sections"]:
        headers, rows = _section_table(section)
        elems.append(Paragraph(section["title"], styles["Heading2"]))
        if not rows:
            elems += [Paragraph("No data.", styles["Normal"]), Spacer(1, 10)]
            continue
        data = [headers] + rows
        col_w = avail_w / len(headers)
        t = Table(data, colWidths=[col_w] * len(headers), repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f5f7")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c8d0d6")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elems += [t, Spacer(1, 14)]

    if detail["module"] != "All":
        elems.append(Paragraph(_SCHEDULE_NOTE, styles["Italic"]))
    doc.build(elems)
    return buf.getvalue()


def _safe_sheet_name(name, used):
    """Excel sheet names: <=31 chars, no []:*?/\\ and unique."""
    clean = "".join(c for c in name if c not in "[]:*?/\\")[:31] or "Sheet"
    base, i = clean, 1
    while clean in used:
        suffix = f" {i}"
        clean = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(clean)
    return clean


def export_excel(report_id, date_from=None, date_to=None):
    """Write each section as its own sheet in an .xlsx workbook. Returns bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    detail = build_detail(report_id, date_from, date_to)
    wb = Workbook()
    used = set()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B242C")

    first = True
    for section in detail["sections"]:
        headers, rows = _section_table(section)
        ws = wb.active if first else wb.create_sheet()
        ws.title = _safe_sheet_name(section["title"], used)
        first = False
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
        for r in rows:
            ws.append(r)
        # autosize columns to their content
        for i, _h in enumerate(headers, start=1):
            width = max([len(str(headers[i - 1]))] + [len(str(r[i - 1])) for r in rows]) + 2
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(width, 48)
        ws.freeze_panes = "A2"

    if first:                      # no sections produced a sheet (shouldn't happen)
        wb.active.title = "Report"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    load_modules()                 # lazy-build everything for a standalone smoke test
    print("Reports:", [r["id"] for r in list_reports()])
    for rid in REPORT_IDS:
        d = build_detail(rid)
        secs = [f"{s['type']}:{len(s.get('items', s.get('rows', s.get('data', []))))}" for s in d["sections"]]
        print(f"  {rid:<22} {d['module']:<10} sections={secs}")
    pdf = export_pdf("cross-module-kpi")
    xlsx = export_excel("fleet-maintenance")
    print(f"PDF bytes: {len(pdf)} · XLSX bytes: {len(xlsx)}")
